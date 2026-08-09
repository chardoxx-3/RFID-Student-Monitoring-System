from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
from django.db.models import Q
from datetime import datetime, date, timedelta
import json
import os
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import pandas as pd
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import tempfile
from django.utils.timezone import localtime
from django.contrib.auth import login, authenticate
from django.utils import timezone
from datetime import timedelta
from .password_reset_forms import ForgotPasswordForm, VerifyOTPForm, ResetPasswordForm
from .models import PasswordResetOTP
from .utils import generate_and_send_otp, verify_otp_and_reset_password
from django.contrib.auth import update_session_auth_hash

from .google_drive_oauth import drive_oauth_service
import base64
import logging
from .forms import (
    CourseForm, SectionForm, StudentForm,
    StudentImportForm, SystemSettingsForm,
    EmailSettingsForm, TestEmailForm,
    UserRegistrationForm, ProfileUpdateForm,
    CustomPasswordChangeForm, InstituteForm  # Add InstituteForm here
)
logger = logging.getLogger(__name__)

# Add Institute to the models import
from .models import (
    Course, Section, Student,
    AttendanceRecord, SystemSettings,
    EmailSettings, SystemLog, Institute  # Add Institute here
)
logger = logging.getLogger(__name__)




from .models import (
    Course, Section, Student,
    AttendanceRecord, SystemSettings,
    EmailSettings, SystemLog
)
from .forms import (
    CourseForm, SectionForm, StudentForm,
    StudentImportForm, SystemSettingsForm,
    EmailSettingsForm, TestEmailForm,
    UserRegistrationForm, ProfileUpdateForm,
    CustomPasswordChangeForm
)
from .utils import (
    send_attendance_email, log_system_event,
    process_student_import, get_dashboard_stats
)
import tempfile

def login_view(request):
    if request.user.is_authenticated:
        return redirect('admin_dashboard')
    
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        action = request.POST.get('action', 'login')
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            if 'remember' not in request.POST:
                request.session.set_expiry(0)
            
            # Store the action in session for the next request
            if action == 'register':
                # Clear any previous messages
                storage = messages.get_messages(request)
                storage.used = True
                return redirect('register')
            else:
                return redirect('admin_dashboard')
        else:
            messages.error(request, 'Invalid username or password')
    
    return render(request, 'attendance_app/login.html')


def logout_view(request):
    logout(request)
    return redirect('login')

def register_view(request):
    # If user is already authenticated, log them out first
    if request.user.is_authenticated:
        logout(request)
        messages.info(request, 'Please register a new account or login with existing credentials.')
    
    # Normal registration flow for unauthenticated users
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Registration successful. Please log in.')
            return redirect('login')
    else:
        form = UserRegistrationForm()
    
    return render(request, 'attendance_app/register.html', {'form': form})

# Admin Views
@login_required
def admin_dashboard(request):
    today = date.today()
    
    # Get today's date in readable format
    today_date = today.strftime("%B %d, %Y")
    
    # Count total students
    total_students = Student.objects.count()
    
    # Get today's attendance records
    today_attendance = AttendanceRecord.objects.filter(date=today).select_related(
        'student', 
        'student__section', 
        'student__section__course'
    )
    
    # Count today's attendance records
    today_attendance_count = today_attendance.count()
    
    # Count AM and PM attendance
    am_present_count = today_attendance.filter(morning_in__isnull=False).count()
    pm_present_count = today_attendance.filter(afternoon_in__isnull=False).count()
    
    # Convert to list for sorting
    attendance_list = list(today_attendance)
    
    # Define a function to get the latest time from a record
    def get_latest_time(record):
        times = [
            record.morning_in,
            record.morning_out,
            record.afternoon_in,
            record.afternoon_out
        ]
        # Filter out None values and return the latest time
        valid_times = [t for t in times if t is not None]
        return max(valid_times) if valid_times else time.min
    
    # Sort records by latest time (descending)
    attendance_list.sort(key=lambda r: get_latest_time(r), reverse=True)
    
    # Get recent attendance records (first 12 after sorting)
    recent_attendance = attendance_list[:12]
    
    context = {
        'today_date': today_date,
        'total_students': total_students,
        'today_attendance_count': today_attendance_count,
        'am_present_count': am_present_count,
        'pm_present_count': pm_present_count,
        'recent_attendance': recent_attendance,
    }
    
    return render(request, 'attendance_app/admin/dashboard.html', context)

# In views.py, add institute views

@login_required
def institute_list(request):
    institutes = Institute.objects.all().order_by('code')
    return render(request, 'attendance_app/admin/institutes.html', {'institutes': institutes})

@login_required
def add_institute(request):
    if request.method == 'POST':
        form = InstituteForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Institute added successfully')
            return redirect('institutes')
        else:
            # Add error message for invalid form with custom tags
            error_message = "Please correct the errors below."
            if 'code' in form.errors:
                error_message = f"Code error: {form.errors['code'][0]}"
            elif 'name' in form.errors:
                error_message = f"Name error: {form.errors['name'][0]}"
            messages.error(request, error_message, extra_tags='duration-5000 alert-danger')
    else:
        form = InstituteForm()
    
    institutes = Institute.objects.all().order_by('code')
    return render(request, 'attendance_app/admin/institutes.html', {
        'form': form,
        'institutes': institutes
    })

@login_required
def edit_institute(request):
    if request.method == 'POST':
        institute_id = request.POST.get('institute_id')
        institute = get_object_or_404(Institute, id=institute_id)
        form = InstituteForm(request.POST, instance=institute)
        if form.is_valid():
            form.save()
            messages.success(request, 'Institute updated successfully')
            return redirect('institutes')
        else:
            # Add error message for invalid form with custom tags
            error_message = "Please correct the errors below."
            if 'code' in form.errors:
                error_message = f"Code error: {form.errors['code'][0]}"
            elif 'name' in form.errors:
                error_message = f"Name error: {form.errors['name'][0]}"
            messages.error(request, error_message, extra_tags='duration-5000 alert-danger')
    return redirect('institutes')

@login_required
def delete_institute(request, institute_id):
    institute = get_object_or_404(Institute, id=institute_id)
    
    if request.method == 'POST':
        institute_name = institute.name
        institute.delete()
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True, 
                'message': f'Institute {institute_name} deleted successfully'
            })
            
        messages.success(request, 'Institute deleted successfully')
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': False, 'message': 'Delete failed'})
    
    return redirect('institutes')

@login_required
def get_courses_by_institute(request, institute_id):
    """API endpoint to get courses by institute"""
    courses = Course.objects.filter(institute_id=institute_id).values('id', 'code', 'name')
    return JsonResponse(list(courses), safe=False)

@login_required
def get_institutes(request):
    """API endpoint to get all institutes"""
    institutes = Institute.objects.all().order_by('code').values('id', 'code', 'name')
    return JsonResponse(list(institutes), safe=False)

@login_required
def course_list(request, institute_id):
    institute = get_object_or_404(Institute, id=institute_id)
    courses = Course.objects.filter(institute=institute).order_by('code')
    return render(request, 'attendance_app/admin/courses.html', {
        'institute': institute,
        'courses': courses
    })

@login_required
def add_course(request):
    if request.method == 'POST':
        form = CourseForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Course added successfully')
            # Redirect back to the institute's courses page
            institute_id = form.cleaned_data['institute'].id
            return redirect('view_courses', institute_id=institute_id)
        else:
            # Add error message for invalid form with custom tags
            error_message = "Please correct the errors below."
            if 'code' in form.errors:
                error_message = f"Code error: {form.errors['code'][0]}"
            elif 'name' in form.errors:
                error_message = f"Name error: {form.errors['name'][0]}"
            elif 'institute' in form.errors:
                error_message = f"Institute error: {form.errors['institute'][0]}"
            messages.error(request, error_message, extra_tags='duration-5000 alert-danger')
    else:
        # Pre-fill the institute if coming from an institute page
        institute_id = request.GET.get('institute_id')
        if institute_id:
            form = CourseForm(initial={'institute': institute_id})
        else:
            form = CourseForm()
    
    # Get the institute for the template context
    institute_id = request.POST.get('institute') or request.GET.get('institute_id')
    if institute_id:
        institute = get_object_or_404(Institute, id=institute_id)
        courses = Course.objects.filter(institute=institute).order_by('code')
    else:
        institute = None
        courses = Course.objects.none()
    
    return render(request, 'attendance_app/admin/courses.html', {
        'institute': institute,
        'courses': courses,
        'form': form if 'form' in locals() else None
    })

@login_required
def edit_course(request):
    if request.method == 'POST':
        course_id = request.POST.get('course_id')
        course = get_object_or_404(Course, id=course_id)
        form = CourseForm(request.POST, instance=course)
        if form.is_valid():
            updated_course = form.save()
            
            # Return JSON response for AJAX or redirect properly
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Course updated successfully'
                })
            
            messages.success(request, 'Course updated successfully')
            # Redirect back to the institute's courses page
            return redirect('view_courses', institute_id=updated_course.institute.id)
        else:
            # Handle form errors
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': form.errors
                }, status=400)
            
            # Add error message for invalid form with custom tags
            error_message = "Please correct the errors below."
            if 'code' in form.errors:
                error_message = f"Code error: {form.errors['code'][0]}"
            elif 'name' in form.errors:
                error_message = f"Name error: {form.errors['name'][0]}"
            elif 'institute' in form.errors:
                error_message = f"Institute error: {form.errors['institute'][0]}"
            messages.error(request, error_message, extra_tags='duration-5000 alert-danger')
    
    return redirect('view_courses', institute_id=course.institute.id)

@login_required
def delete_course(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    
    if request.method == 'POST':
        course.delete()
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
            
        messages.success(request, 'Course deleted successfully')
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': False})
    
    return redirect('courses')

@login_required
def section_list(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    sections = Section.objects.filter(course=course).order_by('year_level', 'name')
    return render(request, 'attendance_app/admin/sections.html', {
        'course': course,
        'sections': sections
    })

@login_required
def add_section(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    if request.method == 'POST':
        name = request.POST.get('name')
        year_level = request.POST.get('year_level')
        
        # Simple validation
        if not name or not year_level:
            messages.error(request, 'Please fill all fields', extra_tags='duration-5000 alert-danger')
            return redirect('view_sections', course_id=course.id)
        
        try:
            # Check for duplicate section name in the same course
            if Section.objects.filter(course=course, name=name).exists():
                messages.error(request, f'A section with name "{name}" already exists in this course', extra_tags='duration-5000 alert-danger')
                return redirect('view_sections', course_id=course.id)
            
            Section.objects.create(
                course=course,
                name=name,
                year_level=year_level
            )
            messages.success(request, 'Section added successfully')
            return redirect('view_sections', course_id=course.id)
        except Exception as e:
            messages.error(request, f'Error adding section: {str(e)}', extra_tags='duration-5000 alert-danger')
            return redirect('view_sections', course_id=course.id)
    
    # If GET request, just show the sections page
    return redirect('view_sections', course_id=course.id)

@login_required
def edit_section(request, section_id):
    section = get_object_or_404(Section, id=section_id)
    
    if request.method == 'POST':
        name = request.POST.get('name')
        year_level = request.POST.get('year_level')
        
        # Check for duplicate section name in the same course (excluding current section)
        if Section.objects.filter(course=section.course, name=name).exclude(id=section_id).exists():
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': {'name': ['A section with this name already exists in this course.']},
                    'message': f'A section with name "{name}" already exists in this course'
                }, status=400)
            else:
                messages.error(request, f'A section with name "{name}" already exists in this course', extra_tags='duration-5000 alert-danger')
                return redirect('view_sections', course_id=section.course.id)
        
        # Create a mutable copy of the POST data
        post_data = request.POST.copy()
        # Add the course ID to the POST data
        post_data['course'] = section.course.id
        
        form = SectionForm(post_data, instance=section)
        if form.is_valid():
            updated_section = form.save()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'section': {
                        'name': updated_section.name,
                        'year_level_display': updated_section.get_year_level_display(),
                        'student_count': updated_section.student_set.count()
                    }
                })
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                # Return specific error messages
                error_message = "Please correct the errors below."
                if 'name' in form.errors:
                    error_message = f"Name error: {form.errors['name'][0]}"
                elif 'year_level' in form.errors:
                    error_message = f"Year level error: {form.errors['year_level'][0]}"
                elif 'course' in form.errors:
                    error_message = f"Course error: {form.errors['course'][0]}"
                
                return JsonResponse({
                    'success': False,
                    'errors': form.errors,
                    'message': error_message
                }, status=400)
            else:
                # For non-AJAX requests, show error message
                error_message = "Please correct the errors below."
                if 'name' in form.errors:
                    error_message = f"Name error: {form.errors['name'][0]}"
                elif 'year_level' in form.errors:
                    error_message = f"Year level error: {form.errors['year_level'][0]}"
                messages.error(request, error_message, extra_tags='duration-5000 alert-danger')
    
    # Handle GET requests
    form = SectionForm(instance=section)
    return render(request, 'attendance_app/admin/sections.html', {
        'course': section.course,
        'form': form,
        'editing': True
    })

@login_required
def delete_section(request, section_id):
    section = get_object_or_404(Section, id=section_id)
    course_id = section.course.id
    
    if request.method == 'POST':
        section.delete()
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
            
        messages.success(request, 'Section deleted successfully')
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': False})
    
    return redirect('view_sections', course_id=course_id)

# In views.py, update the student_list function
# In views.py - update the student_list function
@login_required
def student_list(request):
    students = Student.objects.select_related('section__course__institute')
    
    # Filtering
    search_query = request.GET.get('search', '')
    course_id = request.GET.get('course', '')
    year_level = request.GET.get('year_level', '')
    section_id = request.GET.get('section', '')
    institute_id = request.GET.get('institute', '')
    sort_by = request.GET.get('sort', 'name_asc')  # Default to name ascending
    
    if search_query:
        students = students.filter(
            Q(student_id__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query)
        )
    
    if institute_id:
        students = students.filter(section__course__institute_id=institute_id)
    
    if course_id:
        students = students.filter(section__course_id=course_id)
    
    if year_level:
        students = students.filter(year_level=year_level)
    
    if section_id:
        students = students.filter(section_id=section_id)
    
    # Sorting - UPDATED WITH DATE SORTING
    if sort_by == 'name_asc':
        students = students.order_by('last_name', 'first_name')
    elif sort_by == 'name_desc':
        students = students.order_by('-last_name', '-first_name')
    elif sort_by == 'date_created_asc':
        students = students.order_by('created_at')
    elif sort_by == 'date_created_desc':
        students = students.order_by('-created_at')
    elif sort_by == 'date_modified_asc':
        students = students.order_by('updated_at')
    elif sort_by == 'date_modified_desc':
        students = students.order_by('-updated_at')
    # Keep existing date sorting for backward compatibility
    elif sort_by == 'date_asc':
        students = students.order_by('id')  # Using ID as proxy for creation date
    elif sort_by == 'date_desc':
        students = students.order_by('-id')  # Using ID as proxy for creation date
    
    # Pagination
    paginator = Paginator(students, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    all_sections = Section.objects.all().order_by('course__name', 'year_level', 'name')
    all_courses = Course.objects.all().order_by('name')
    all_institutes = Institute.objects.all().order_by('name')
    
    return render(request, 'attendance_app/admin/students.html', {
        'students': page_obj,
        'all_sections': all_sections,
        'all_courses': all_courses,
        'all_institutes': all_institutes,
        'search_query': search_query,
        'selected_institute': institute_id,
        'selected_course': course_id,
        'selected_year': year_level,
        'selected_section': section_id,
        'selected_sort': sort_by  # Add selected sort to context
    })

@login_required
def add_student(request):
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES)
        if form.is_valid():
            student = form.save()
            messages.success(request, f'Student {student.full_name} added successfully')
            return redirect('students')
        else:
            # Enhanced error message for invalid form
            error_message = "Please correct the errors below."
            if 'student_id' in form.errors:
                error_message = f"Student ID error: {form.errors['student_id'][0]}"
            elif 'first_name' in form.errors:
                error_message = f"First name error: {form.errors['first_name'][0]}"
            elif 'last_name' in form.errors:
                error_message = f"Last name error: {form.errors['last_name'][0]}"
            elif 'section' in form.errors:
                error_message = f"Section error: {form.errors['section'][0]}"
            elif 'year_level' in form.errors:
                error_message = f"Year level error: {form.errors['year_level'][0]}"
            elif 'rfid_tag' in form.errors:
                error_message = f"RFID tag error: {form.errors['rfid_tag'][0]}"
            elif 'photo' in form.errors:
                error_message = f"Photo error: {form.errors['photo'][0]}"
            
            messages.error(request, error_message, extra_tags='duration-5000 alert-danger')
    
    all_sections = Section.objects.all().order_by('name')
    return render(request, 'attendance_app/admin/students.html', {
        'form': form if 'form' in locals() else StudentForm(),
        'all_sections': all_sections
    })

@login_required
def edit_student(request):
    if request.method == 'POST':
        student_db_id = request.POST.get('student_db_id')
        student = get_object_or_404(Student, id=student_db_id)
        
        form = StudentForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, f'Student {student.full_name} updated successfully')
            return redirect('students')
        else:
            # Enhanced error message for invalid form
            error_message = "Please correct the errors below."
            if 'student_id' in form.errors:
                error_message = f"Student ID error: {form.errors['student_id'][0]}"
            elif 'first_name' in form.errors:
                error_message = f"First name error: {form.errors['first_name'][0]}"
            elif 'last_name' in form.errors:
                error_message = f"Last name error: {form.errors['last_name'][0]}"
            elif 'section' in form.errors:
                error_message = f"Section error: {form.errors['section'][0]}"
            elif 'year_level' in form.errors:
                error_message = f"Year level error: {form.errors['year_level'][0]}"
            elif 'rfid_tag' in form.errors:
                error_message = f"RFID tag error: {form.errors['rfid_tag'][0]}"
            elif 'photo' in form.errors:
                error_message = f"Photo error: {form.errors['photo'][0]}"
            
            messages.error(request, error_message, extra_tags='duration-5000 alert-danger')
    
    return redirect('students')

@login_required
def delete_student(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    if request.method == 'POST':
        student.delete()
        messages.success(request, 'Student deleted successfully')
    return redirect('students')

@login_required
def import_students(request):
    if request.method == 'POST':
        form = StudentImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                excel_file = request.FILES['excel_file']
                
                # Check file extension
                if not excel_file.name.endswith(('.xlsx', '.xls')):
                    messages.error(request, "Invalid file type. Please upload an Excel file (.xlsx or .xls)")
                    return redirect('students')
                
                # Use pandas to read the Excel file
                try:
                    dtype_spec = {
                        'student_id': str,
                        'rfid_tag': str  # Treat RFID as string to preserve leading zeros
                    }
                    df = pd.read_excel(excel_file, dtype=dtype_spec)
                    
                    # DEBUG: Print column names and first few rows
                    print("=== DEBUG: Excel File Analysis ===")
                    print("Columns in Excel:", df.columns.tolist())
                    print("RFID column sample:", df['rfid_tag'].head(10).tolist() if 'rfid_tag' in df.columns else "No RFID column")
                    print("================================")
                    
                except Exception as e:
                    messages.error(request, f"Error reading Excel file: {str(e)}")
                    return redirect('students')
                
                # Process the data
                result = {
                    'success': True,
                    'imported': 0,
                    'updated': 0,
                    'skipped': 0,
                    'errors': []
                }
                
                # Map your ACTUAL column names to what the code expects
                column_mapping = {
                    'department': 'institute',
                    'program': 'course'
                }
                
                # Rename columns to match what the backend logic expects
                df = df.rename(columns=column_mapping)
                
                required_columns = [
                    'student_id', 'first_name', 'last_name', 'birthday',
                    'gender', 'year_level', 'institute', 'course', 'section', 'email'
                ]
                
                # Check if all required columns are present
                missing_columns = [col for col in required_columns if col not in df.columns]
                if missing_columns:
                    result['success'] = False
                    result['error'] = f"Missing required columns: {', '.join(missing_columns)}. Found columns: {', '.join(df.columns.tolist())}"
                    messages.error(request, result['error'])
                    return redirect('students')
                
                update_existing = form.cleaned_data['update_existing']
                
                for index, row in df.iterrows():
                    try:
                        # DEBUG: Print current row data
                        print(f"=== Processing row {index+2} ===")
                        print(f"Student ID: {row['student_id']}")
                        print(f"RFID Tag (raw): {repr(row.get('rfid_tag', ''))}")
                        
                        # GET OR CREATE INSTITUTE
                        institute_name = str(row['institute']).strip()

                        # Extract code and name if format is "CODE - Name"
                        if ' - ' in institute_name:
                            code_part, name_part = institute_name.split(' - ', 1)
                            institute_code = code_part.strip()
                            institute_display_name = name_part.strip()
                        else:
                            # If no code format, use the whole string as name and generate a code
                            institute_code = institute_name[:10].upper().replace(' ', '')  # Generate code from first 10 chars
                            institute_display_name = institute_name

                        # Try to find by code first, then by name
                        institute = Institute.objects.filter(
                            Q(code__iexact=institute_code) | Q(name__icontains=institute_display_name)
                        ).first()

                        if not institute:
                            # Create new institute if not found
                            try:
                                institute = Institute.objects.create(
                                    code=institute_code,
                                    name=institute_display_name
                                )
                                print(f"CREATED NEW INSTITUTE: {institute.code} - {institute.name}")
                            except Exception as e:
                                error_msg = f"Row {index+2}: Failed to create institute '{institute_name}': {str(e)}"
                                result['errors'].append(error_msg)
                                result['skipped'] += 1
                                print(f"ERROR: {error_msg}")
                                continue
                        
                        # GET OR CREATE COURSE
                        course_name = str(row['course']).strip()

                        # Extract code and name if format is "CODE - Name"
                        if ' - ' in course_name:
                            course_code_part, course_name_part = course_name.split(' - ', 1)
                            course_code = course_code_part.strip()
                            course_display_name = course_name_part.strip()
                        else:
                            # If no code format, use the whole string as name and generate a code
                            course_code = course_name[:8].upper().replace(' ', '')  # Generate code from first 8 chars
                            course_display_name = course_name

                        course = Course.objects.filter(
                            Q(code__iexact=course_code) | Q(name__icontains=course_display_name),
                            institute=institute
                        ).first()

                        if not course:
                            # Create new course if not found
                            try:
                                course = Course.objects.create(
                                    code=course_code,
                                    name=course_display_name,
                                    institute=institute
                                )
                                print(f"CREATED NEW COURSE: {course.code} - {course.name} in {institute.name}")
                            except Exception as e:
                                error_msg = f"Row {index+2}: Failed to create course '{course_name}' in institute '{institute.name}': {str(e)}"
                                result['errors'].append(error_msg)
                                result['skipped'] += 1
                                print(f"ERROR: {error_msg}")
                                continue
                        
                        # Get or create the section
                        section_name = str(row['section']).strip()
                        year_level = int(row['year_level'])
                        
                        section, created = Section.objects.get_or_create(
                            name=section_name,
                            course=course,
                            year_level=year_level,
                            defaults={'name': section_name}
                        )
                        
                        # FIX: Proper RFID tag handling
                        rfid_value = row.get('rfid_tag', '')
                        
                        # Handle different cases for empty RFID tags
                        if pd.isna(rfid_value) or rfid_value == '' or str(rfid_value).lower() in ['nan', 'null', 'none', '']:
                            rfid_final = None  # Use None for database NULL
                        else:
                            rfid_final = str(rfid_value).strip()
                            # If it's just 'nan' as string, treat as empty
                            if rfid_final.lower() == 'nan':
                                rfid_final = None
                        
                        print(f"RFID Tag (processed): {repr(rfid_final)}")
                        
                        # Check for duplicate RFID if not None
                        if rfid_final and Student.objects.filter(rfid_tag=rfid_final).exists():
                            # If updating existing, we'll handle it in update_or_create
                            # If not updating, skip
                            if not update_existing:
                                existing_student = Student.objects.filter(rfid_tag=rfid_final).first()
                                error_msg = f"Row {index+2}: RFID tag '{rfid_final}' already assigned to {existing_student.full_name}"
                                result['errors'].append(error_msg)
                                result['skipped'] += 1
                                print(f"ERROR: {error_msg}")
                                continue
                        
                        # Prepare student data
                        student_data = {
                            'student_id': str(row['student_id']).strip(),
                            'first_name': str(row['first_name']).strip(),
                            'last_name': str(row['last_name']).strip(),
                            'birthday': row['birthday'] if pd.notna(row['birthday']) else None,
                            'gender': row['gender'][0].upper() if pd.notna(row['gender']) else 'O',
                            'year_level': year_level,
                            'section': section,
                            'email': row['email'] if pd.notna(row['email']) else '',
                            'address': row.get('address', ''),
                            'guardian_name': row.get('guardian_name', ''),
                            'guardian_email': row.get('guardian_email', ''),
                            'rfid_tag': rfid_final,  # Use the properly processed RFID value
                        }
                        
                        print(f"Student data prepared: {student_data['first_name']} {student_data['last_name']}")
                        
                        if update_existing:
                            # For update_or_create, we need to handle the case where we might be updating
                            # a student that has the same RFID as another student
                            existing_with_rfid = None
                            if rfid_final:
                                existing_with_rfid = Student.objects.filter(rfid_tag=rfid_final).exclude(student_id=student_data['student_id']).first()
                            
                            if existing_with_rfid:
                                # RFID is assigned to a different student, skip this row
                                error_msg = f"Row {index+2}: RFID tag '{rfid_final}' already assigned to {existing_with_rfid.full_name} (ID: {existing_with_rfid.student_id})"
                                result['errors'].append(error_msg)
                                result['skipped'] += 1
                                print(f"ERROR: {error_msg}")
                                continue
                            
                            student, created = Student.objects.update_or_create(
                                student_id=student_data['student_id'],
                                defaults=student_data
                            )
                            if created:
                                result['imported'] += 1
                                print(f"SUCCESS: Imported new student")
                            else:
                                result['updated'] += 1
                                print(f"SUCCESS: Updated existing student")
                        else:
                            if not Student.objects.filter(student_id=student_data['student_id']).exists():
                                # Check for duplicate RFID for new students
                                if rfid_final and Student.objects.filter(rfid_tag=rfid_final).exists():
                                    existing_student = Student.objects.filter(rfid_tag=rfid_final).first()
                                    error_msg = f"Row {index+2}: RFID tag '{rfid_final}' already assigned to {existing_student.full_name}"
                                    result['errors'].append(error_msg)
                                    result['skipped'] += 1
                                    print(f"ERROR: {error_msg}")
                                    continue
                                
                                Student.objects.create(**student_data)
                                result['imported'] += 1
                                print(f"SUCCESS: Created new student")
                            else:
                                result['skipped'] += 1
                                error_msg = f"Row {index+2}: Student ID {student_data['student_id']} already exists"
                                result['errors'].append(error_msg)
                                print(f"SKIPPED: {error_msg}")
                                
                    except Exception as e:
                        result['skipped'] += 1
                        error_msg = f"Row {index+2}: {str(e)}"
                        result['errors'].append(error_msg)
                        print(f"EXCEPTION: {error_msg}")
                        import traceback
                        traceback.print_exc()
                
                print("=== IMPORT SUMMARY ===")
                print(f"Imported: {result['imported']}, Updated: {result['updated']}, Skipped: {result['skipped']}")
                print(f"Errors: {len(result['errors'])}")
                for error in result['errors'][:5]:
                    print(f" - {error}")
                print("=====================")
                
                if result['success']:
                    msg = f"Import completed: {result['imported']} imported, {result['updated']} updated, {result['skipped']} skipped"
                    if result['errors']:
                        msg += f". {len(result['errors'])} errors occurred."
                        request.session['import_errors'] = result['errors'][:10]
                    messages.success(request, msg)
                else:
                    messages.error(request, f"Import failed: {result['error']}")
                
            except Exception as e:
                messages.error(request, f"Import failed: {str(e)}")
                import traceback
                traceback.print_exc()
            
            return redirect('students')
    
    messages.error(request, "Invalid request")
    return redirect('students')



@login_required
def export_students(request):
    # Get filtered students based on the same filters used in student_list
    students = Student.objects.select_related('section__course__institute').order_by('last_name', 'first_name')
    
    # Apply the same filters as in student_list view
    search_query = request.GET.get('search', '')
    institute_id = request.GET.get('institute', '')
    course_id = request.GET.get('course', '')
    year_level = request.GET.get('year_level', '')
    section_id = request.GET.get('section', '')
    
    if search_query:
        students = students.filter(
            Q(student_id__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query)
        )
    
    if institute_id:
        students = students.filter(section__course__institute_id=institute_id)
    
    if course_id:
        students = students.filter(section__course_id=course_id)
    
    if year_level:
        students = students.filter(year_level=year_level)
    
    if section_id:
        students = students.filter(section_id=section_id)
    
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    
    # Add headers (should match your import format) - UPDATED to include institute
    headers = [
        "student_id", "rfid_tag", "first_name", "last_name", 
        "birthday", "gender", "year_level", "institute", "course", "section",  # Added institute before course
        "email", "address", "guardian_name", "guardian_email"
    ]
    
    # Write headers with bold style
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}1"]
        cell.value = header
        cell.font = Font(bold=True)
    
    # Write student data - UPDATED to include institute
    for row_num, student in enumerate(students, 2):
        ws.cell(row=row_num, column=1, value=student.student_id)
        ws.cell(row=row_num, column=2, value=student.rfid_tag or "")
        ws.cell(row=row_num, column=3, value=student.first_name)
        ws.cell(row=row_num, column=4, value=student.last_name)
        ws.cell(row=row_num, column=5, value=student.birthday.strftime('%Y-%m-%d') if student.birthday else "")
        ws.cell(row=row_num, column=6, value=student.get_gender_display())
        ws.cell(row=row_num, column=7, value=student.year_level)
        # NEW: Add institute code and name
        ws.cell(row=row_num, column=8, value=f"{student.section.course.institute.code} - {student.section.course.institute.name}" if student.section and student.section.course and student.section.course.institute else "")
        # Updated course column (now column 9)
        ws.cell(row=row_num, column=9, value=f"{student.section.course.code} - {student.section.course.name}" if student.section and student.section.course else "")
        # Updated section column (now column 10)
        ws.cell(row=row_num, column=10, value=student.section.name if student.section else "")
        # Updated remaining columns
        ws.cell(row=row_num, column=11, value=student.email or "")
        ws.cell(row=row_num, column=12, value=student.address or "")
        ws.cell(row=row_num, column=13, value=student.guardian_name or "")
        ws.cell(row=row_num, column=14, value=student.guardian_email or "")
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=students_export.xlsx'
    wb.save(response)
    
    return response

@login_required
def attendance_reports(request):
    # Default to show all records (remove the 7-day limit)
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    institute_id = request.GET.get('institute')  # Added institute filter
    section_id = request.GET.get('section')
    course_id = request.GET.get('course')
    year_level = request.GET.get('year_level')
    search_query = request.GET.get('search', '')
    
    # If no dates provided, show all records
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        # Set to a very old date to get all records
        start_date = date(2020, 1, 1)  # Or your system's start date
    
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        # Set to future date to get all records up to today
        end_date = date.today() + timedelta(days=365)  # Include future dates if any
    
    # Get all records first
    records = AttendanceRecord.objects.select_related(
        'student', 
        'student__section', 
        'student__section__course',
        'student__section__course__institute'
    ).filter(
        date__range=[start_date, end_date]
    )
    
    # Apply filters
    if institute_id:
        records = records.filter(student__section__course__institute_id=institute_id)
    
    if section_id:
        records = records.filter(student__section_id=section_id)
    
    if course_id:
        records = records.filter(student__section__course_id=course_id)
    
    if year_level:
        records = records.filter(student__year_level=year_level)
    
    if search_query:
        records = records.filter(
            Q(student__student_id__icontains=search_query) |
            Q(student__first_name__icontains=search_query) |
            Q(student__last_name__icontains=search_query)
        )
    
    # Convert to list for custom sorting by latest time
    records_list = list(records)
    
    # Define a function to get the latest time from a record
    def get_latest_time(record):
        times = [
            record.morning_in,
            record.morning_out,
            record.afternoon_in,
            record.afternoon_out
        ]
        # Filter out None values and return the latest time
        valid_times = [t for t in times if t is not None]
        return max(valid_times) if valid_times else datetime.min.time()
    
    # Sort records by latest time (descending) and then by date (descending)
    records_list.sort(key=lambda r: (r.date, get_latest_time(r)), reverse=True)
    
    # Get all institutes, courses and sections for filter dropdowns
    all_institutes = Institute.objects.all().order_by('code')
    all_courses = Course.objects.all().order_by('code')
    all_sections = Section.objects.all().order_by('name')
    
    # Pagination - 20 records per page
    paginator = Paginator(records_list, 20)
    page_number = request.GET.get('page')
    
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page
        page_number = 1
        page_obj = paginator.page(page_number)
    except EmptyPage:
        # If page is out of range, deliver last page
        page_obj = paginator.page(paginator.num_pages)
    
    return render(request, 'attendance_app/admin/reports.html', {
        'attendance_records': page_obj,
        'all_institutes': all_institutes,  # Added
        'all_courses': all_courses,
        'all_sections': all_sections,
        'start_date': start_date.strftime('%Y-%m-%d') if start_date else '',
        'end_date': end_date.strftime('%Y-%m-%d') if end_date else '',
        'selected_institute': institute_id,  # Added
        'selected_course': course_id,
        'selected_year': year_level,
        'selected_section': section_id,
        'search_query': search_query,
        'paginator': paginator,
        'page_obj': page_obj
    })

# Settings Views
@login_required
def settings_view(request):
    return redirect('profile_settings')

@login_required
def profile_settings(request):
    if request.method == 'POST':
        form = ProfileUpdateForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Profile updated successfully')
            return redirect('profile_settings')
    else:
        form = ProfileUpdateForm(instance=request.user)
    
    return render(request, 'attendance_app/admin/settings/profile.html', {'form': form})

@login_required
def email_settings(request):
    email_settings, created = EmailSettings.objects.get_or_create(pk=1)
    
    if request.method == 'POST':
        form = EmailSettingsForm(request.POST, instance=email_settings)
        if form.is_valid():
            form.save()
            messages.success(request, 'Email settings updated successfully')
            return redirect('email_settings')
        else:
            # Add this to see form errors in console
            print(form.errors)
            messages.error(request, 'Please correct the errors below')
    else:
        form = EmailSettingsForm(instance=email_settings)
    
    return render(request, 'attendance_app/admin/settings/email_setup.html', {
        'email_settings': email_settings,
        'form': form,
        'test_form': TestEmailForm()
    })



@login_required
def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Important to keep the user logged in
            messages.success(request, 'Your password was successfully updated!')
            return redirect('profile_settings')
    else:
        form = PasswordChangeForm(request.user)
    
    return render(request, 'attendance_app/admin/settings/profile.html', {
        'password_form': form
    })

@login_required
def send_test_email(request):
    if request.method == 'POST':
        form = TestEmailForm(request.POST)
        if form.is_valid():
            # Create a test student
            student = Student(
                first_name="Test",
                last_name="Student",
                section=Section.objects.first(),
                year_level=1,
                student_id="TEST001",
                guardian_email=form.cleaned_data['test_email']
            )
            
            # Create a proper attendance record using the actual model fields
            today = date.today()
            now = datetime.now().time()
            is_morning = now.hour < 12
            
            record = AttendanceRecord(
                student=student,
                date=today,
            )
            
            # Set the appropriate time field based on email type and time of day
            if form.cleaned_data['test_type'] == 'time_in':
                if is_morning:
                    record.morning_in = now
                else:
                    record.afternoon_in = now
            else:  # time_out
                if is_morning:
                    record.morning_out = now
                else:
                    record.afternoon_out = now
            
            # Send test email
            success = send_attendance_email(
                student, 
                record, 
                form.cleaned_data['test_type']
            )
            
            if success:
                messages.success(request, 'Test email sent successfully')
            else:
                messages.error(request, 'Failed to send test email')
    
    return redirect('email_settings')


# In your system_settings view, update the context:
@login_required
def system_settings(request):
    system_settings, created = SystemSettings.objects.get_or_create(pk=1)
    
    if request.method == 'POST':
        form = SystemSettingsForm(request.POST, request.FILES, instance=system_settings)
        if form.is_valid():
            form.save()
            messages.success(request, 'System settings updated successfully')
            return redirect('system_settings')
    else:
        form = SystemSettingsForm(instance=system_settings)
    
    # Get all logs ordered by timestamp (newest first)
    logs_list = SystemLog.objects.all().order_by('-timestamp')
    
    # Pagination - show 20 logs per page
    paginator = Paginator(logs_list, 10)
    page = request.GET.get('page')
    
    try:
        system_logs = paginator.page(page)
    except PageNotAnInteger:
        system_logs = paginator.page(1)
    except EmptyPage:
        system_logs = paginator.page(paginator.num_pages)
    
    # Check Google Drive authentication status
    drive_authenticated = drive_oauth_service.is_authenticated()
    
    return render(request, 'attendance_app/admin/settings/system_info.html', {
        'form': form,
        'system_logs': system_logs,
        'drive_authenticated': drive_authenticated  # Add this line
    })

# Kiosk Views
def attendance_kiosk(request):
    return render(request, 'attendance_app/kiosk.html')

@csrf_exempt
def log_attendance(request, rfid_tag):
    """Enhanced attendance logging with offline support and Google Drive integration"""
    logger = logging.getLogger(__name__)
    
    # Start request logging
    logger.info(
        f"Attendance log request - Method: {request.method}, "
        f"RFID: {rfid_tag}, "
        f"Headers: {dict(request.headers)}"
    )

    if request.method != 'POST':
        logger.error("Invalid request method - only POST allowed")
        return JsonResponse({
            'success': False,
            'error': 'Invalid request method',
            'debug': {
                'received_method': request.method,
                'expected_method': 'POST'
            }
        })

    try:
        # Normalize RFID tag
        original_tag = rfid_tag
        rfid_tag = rfid_tag.strip()
        
        if not rfid_tag:
            logger.error("Empty RFID tag received")
            return JsonResponse({
                'success': False,
                'error': 'Empty RFID tag',
                'debug': {
                    'original_tag': original_tag,
                    'normalized_tag': rfid_tag
                }
            })

        logger.info(f"Processing RFID tag (normalized): '{rfid_tag}'")

        # Find student with case-insensitive match
        student = Student.objects.filter(rfid_tag__iexact=rfid_tag).select_related('section__course').first()
        
        if not student:
            similar_tags = Student.objects.filter(
                rfid_tag__icontains=rfid_tag
            ).values_list('rfid_tag', flat=True)[:5]
            
            logger.error(
                f"Student not found - Tag: '{rfid_tag}'. "
                f"Similar tags: {list(similar_tags)}"
            )
            
            return JsonResponse({
                'success': False,
                'error': 'Student not found',
                'debug': {
                    'received_tag': rfid_tag,
                    'similar_tags': list(similar_tags),
                    'suggestion': 'Check for leading/trailing spaces or case differences'
                }
            })

        logger.info(f"Found student: {student.full_name} (ID: {student.student_id})")

        # Process time and attendance period
        today = date.today()
        now = datetime.now().time()
        current_hour = now.hour
        is_morning = current_hour < 12  # Before 12 PM is morning
        
        logger.info(
            f"Time: {now} (Hour: {current_hour}) - "
            f"Period: {'Morning' if is_morning else 'Afternoon'}"
        )

        # Handle photo capture
        photo_file = None
        photo_data = None
        photo_name = None
        if 'photo' in request.FILES:
            photo_file = request.FILES['photo']
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            photo_name = f"{student.student_id}_{'morning' if is_morning else 'afternoon'}_{timestamp}.jpg"
            photo_file.name = photo_name
            logger.info(f"Photo captured: {photo_name}")
            
            # Read photo data for storage
            photo_data = photo_file.read()
            photo_file.seek(0)  # Reset file pointer for saving to model

        # Get or create attendance record
        record, created = AttendanceRecord.objects.get_or_create(
            student=student,
            date=today,
            defaults={
                'morning_in': None,
                'morning_out': None,
                'afternoon_in': None,
                'afternoon_out': None,
                'morning_in_photo_url': None,
                'morning_out_photo_url': None,
                'afternoon_in_photo_url': None,
                'afternoon_out_photo_url': None,
                'email_sent': False,
                'email_attempted': False
            }
        )
        
        logger.info(
            f"Attendance record {'created' if created else 'updated'} - "
            f"Existing times - "
            f"Morning: {record.morning_in}|{record.morning_out} "
            f"Afternoon: {record.afternoon_in}|{record.afternoon_out}"
        )

        # Determine which field to update
        time_field = None
        email_type = None
        message = None
        
        if is_morning:
            if not record.morning_in:
                time_field = 'morning_in'
                email_type = 'time_in'
                message = f"Morning Time In at {now.strftime('%I:%M %p')}"
            elif not record.morning_out:
                time_field = 'morning_out'
                email_type = 'time_out'
                message = f"Morning Time Out at {now.strftime('%I:%M %p')}"
        else:
            if not record.afternoon_in:
                time_field = 'afternoon_in'
                email_type = 'time_in'
                message = f"Afternoon Time In at {now.strftime('%I:%M %p')}"
            elif not record.afternoon_out:
                time_field = 'afternoon_out'
                email_type = 'time_out'
                message = f"Afternoon Time Out at {now.strftime('%I:%M %p')}"

        # If no field to update (all times logged)
        if not time_field:
            logger.warning("All attendance times already logged for today")
            return JsonResponse({
                'success': True,
                'message': 'Attendance already completed today',
                'student_name': student.full_name,
                'student_id': student.student_id,
                'attendance': {
                    'morning_in': record.morning_in.strftime('%I:%M %p') if record.morning_in else None,
                    'morning_out': record.morning_out.strftime('%I:%M %p') if record.morning_out else None,
                    'afternoon_in': record.afternoon_in.strftime('%I:%M %p') if record.afternoon_in else None,
                    'afternoon_out': record.afternoon_out.strftime('%I:%M %p') if record.afternoon_out else None
                }
            })

        # Update the record
        setattr(record, time_field, now)
        
        # Check if we're in offline mode
        is_offline = request.headers.get('X-Offline-Mode', 'false') == 'true'
        
        if is_offline:
            logger.info("System is offline - marking email as pending")
            record.email_sent = False
            record.email_attempted = False
        else:
            record.email_attempted = True

        # Save the record to database first
        record.save()
        logger.info(f"Updated {time_field} successfully and saved to database")
        
        # Upload photo to Google Drive in background (if online)
        if photo_data and photo_name and not is_offline:
            try:
                # Get Google Drive folder ID from settings
                drive_folder_id = getattr(settings, 'GOOGLE_DRIVE_FOLDER_ID', None)
                
                # Define callback function to update record with photo URL
                def upload_callback(result_url):
                    if result_url:
                        try:
                            # Update record with photo URL
                            photo_url_field = f"{time_field}_photo_url"
                            setattr(record, photo_url_field, result_url)
                            record.save()
                            logger.info(f"Photo uploaded to Google Drive: {result_url}")
                        except Exception as e:
                            logger.error(f"Error updating record with photo URL: {e}")
                    else:
                        logger.warning("Failed to upload photo to Google Drive")
                
                # Queue the upload (non-blocking)
                drive_oauth_service.queue_upload(
                    photo_data, 
                    photo_name, 
                    drive_folder_id, 
                    upload_callback
                )
                
                logger.info("Photo upload queued for background processing")
                    
            except Exception as e:
                logger.error(f"Error queuing photo upload to Google Drive: {e}")
        
        # Handle offline photo storage
        elif photo_data and photo_name and is_offline:
            try:
                # Save photo to local storage
                from django.core.files.base import ContentFile
                from django.core.files.storage import default_storage
                import uuid
                
                # Create attendance_photos directory if it doesn't exist
                photos_dir = os.path.join(settings.MEDIA_ROOT, 'attendance_photos')
                os.makedirs(photos_dir, exist_ok=True)
                
                # Generate a unique filename
                unique_filename = f"{uuid.uuid4().hex}_{photo_name}"
                file_path = f"attendance_photos/{unique_filename}"
                
                # Save to local storage
                saved_path = default_storage.save(file_path, ContentFile(photo_data))
                local_photo_url = default_storage.url(saved_path)
                
                # Update record with local photo URL
                photo_url_field = f"{time_field}_photo_url"
                setattr(record, photo_url_field, local_photo_url)
                record.save()
                
                logger.info(f"Photo saved locally for offline mode: {local_photo_url}")
                
            except Exception as e:
                logger.error(f"Error saving photo locally for offline mode: {e}")

        # Send email notification if applicable (only if online)
        if email_type and student.guardian_email and not is_offline:
            logger.info(f"Preparing to send {email_type} email to {student.guardian_email}")
            
            try:
                # Send email with the actual photo file (not URL)
                email_sent = send_attendance_email(
                    student=student,
                    record=record,
                    email_type=email_type,
                    photo_file=photo_file,  # Use the file object directly
                    photo_filename=photo_name
                )
                
                if email_sent:
                    record.email_sent = True
                    record.save()
                    logger.info(f"Email sent successfully to {student.guardian_email}")
                else:
                    logger.warning(f"Failed to send email to {student.guardian_email}")
                    
            except Exception as e:
                logger.error(f"Error sending email notification: {str(e)}")

        # Prepare response data
        response_data = {
            'success': True,
            'student_name': student.full_name,
            'student_id': student.student_id,
            'course': student.section.course.name,
            'section': student.section.name,
            'year_level': student.get_year_level_display(),
            'message': message,
            'period': 'morning' if is_morning else 'afternoon',
            'action': email_type.split('_')[1] if email_type else None,
            'student_photo_url': student.photo.url if student.photo else None,
            'morning_in': record.morning_in.strftime('%I:%M %p') if record.morning_in else None,
            'morning_out': record.morning_out.strftime('%I:%M %p') if record.morning_out else None,
            'afternoon_in': record.afternoon_in.strftime('%I:%M %p') if record.afternoon_in else None,
            'afternoon_out': record.afternoon_out.strftime('%I:%M %p') if record.afternoon_out else None,
            'is_offline': is_offline
        }

        logger.info("Attendance logged successfully")
        return JsonResponse(response_data)

    except Exception as e:
        logger.error(
            f"Unexpected error processing attendance - "
            f"RFID: {rfid_tag}, "
            f"Error: {str(e)}",
            exc_info=True
        )
        return JsonResponse({
            'success': False,
            'error': 'System error processing attendance',
            'debug': {
                'received_tag': rfid_tag,
                'error_type': type(e).__name__,
                'error_details': str(e)
            }
        })
    

@csrf_exempt
def log_attendance_quick(request, rfid_tag):
    """Quick attendance logging without photo/email processing"""
    try:
        rfid_tag = rfid_tag.strip()
        student = Student.objects.filter(rfid_tag__iexact=rfid_tag).select_related('section__course').first()
        
        if not student:
            return JsonResponse({
                'success': False,
                'error': 'Student not found'
            })

        today = date.today()
        now = datetime.now().time()
        is_morning = now.hour < 12

        # Get or create record
        record, _ = AttendanceRecord.objects.get_or_create(
            student=student,
            date=today,
            defaults={
                'morning_in': None,
                'morning_out': None,
                'afternoon_in': None,
                'afternoon_out': None
            }
        )

        # Determine which action would be taken
        message = None
        if is_morning:
            if not record.morning_in:
                message = f"Morning Time In at {now.strftime('%I:%M %p')}"
            elif not record.morning_out:
                message = f"Morning Time Out at {now.strftime('%I:%M %p')}"
            else:
                message = "Morning session already completed"
        else:
            if not record.afternoon_in:
                message = f"Afternoon Time In at {now.strftime('%I:%M %p')}"
            elif not record.afternoon_out:
                message = f"Afternoon Time Out at {now.strftime('%I:%M %p')}"
            else:
                message = "Afternoon session already completed"

        return JsonResponse({
            'success': True,
            'student_name': student.full_name,
            'student_id': student.student_id,
            'course': student.section.course.name,
            'section': student.section.name,
            'message': message,
            'student_photo_url': student.photo.url if student.photo else None,
            'can_log': not (
                (is_morning and record.morning_in and record.morning_out) or
                (not is_morning and record.afternoon_in and record.afternoon_out)
            )
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': 'System error processing attendance'
        })

@csrf_exempt
def rfid_student_map(request):
    """Returns a mapping of all RFID tags to basic student data"""
    students = Student.objects.select_related('section__course__institute').exclude(rfid_tag__isnull=True)
    
    data = {
        student.rfid_tag: {
            'full_name': student.full_name,
            'student_id': student.student_id,
            'course': student.section.course.name,
            'section': student.section.name,
            'institute': student.section.course.institute.name if student.section.course.institute else None,  # Add institute
            'photo_url': student.photo.url if student.photo else None
        }
        for student in students
    }
    
    return JsonResponse(data)

# Add to views.py
@csrf_exempt
def debug_rfid_tags(request):
    from django.http import JsonResponse
    tags = list(Student.objects.exclude(rfid_tag__isnull=True).values_list('rfid_tag', flat=True))
    return JsonResponse({'rfid_tags': tags})

@login_required
def delete_all_students(request):
    if request.method == 'POST':
        try:
            # Delete all students
            count, _ = Student.objects.all().delete()
            messages.success(request, f'Successfully deleted {count} students')
        except Exception as e:
            messages.error(request, f'Error deleting students: {str(e)}')
    
    return redirect('students')



@csrf_exempt
def kiosk_attendance_records(request):
    """Returns today's attendance records for the kiosk display along with all students"""
    today = date.today()
    
    # Get all students with RFID tags (exclude null values)
    students_with_rfid = Student.objects.exclude(rfid_tag__isnull=True).select_related(
        'section__course'
    )
    
    # Get today's attendance records
    attendance_records = AttendanceRecord.objects.filter(date=today).select_related(
        'student'
    )
    
    # Convert to dictionary for easy lookup
    attendance_dict = {
        record.student.student_id: {
            'morning_in': record.morning_in.strftime('%H:%M:%S') if record.morning_in else None,
            'morning_out': record.morning_out.strftime('%H:%M:%S') if record.morning_out else None,
            'afternoon_in': record.afternoon_in.strftime('%H:%M:%S') if record.afternoon_in else None,
            'afternoon_out': record.afternoon_out.strftime('%H:%M:%S') if record.afternoon_out else None,
        }
        for record in attendance_records
    }
    
    # Prepare response data
    data = []
    for student in students_with_rfid:
        photo_url = student.photo.url if student.photo else None
        record = attendance_dict.get(student.student_id, {})
        
        data.append({
            'student_id': student.student_id,
            'student_name': f"{student.first_name} {student.last_name}",
            'student_photo_url': photo_url,
            'morning_in': record.get('morning_in'),
            'morning_out': record.get('morning_out'),
            'afternoon_in': record.get('afternoon_in'),
            'afternoon_out': record.get('afternoon_out'),
            'course': student.section.course.name if student.section else None,
            'section': student.section.name if student.section else None,
            'rfid_tag': student.rfid_tag
        })
    
    return JsonResponse({'records': data})


@login_required
def attendance_chart_data(request):
    # Get the same filters as the reports view
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    section_id = request.GET.get('section')
    course_id = request.GET.get('course')
    year_level = request.GET.get('year_level')
    search_query = request.GET.get('search', '')
    
    # For charts, if no dates provided, use a reasonable default range
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        # Default to last 30 days for charts if no date filter
        start_date = date.today() - timedelta(days=30)
    
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        end_date = date.today()
    
    # Get records for the date range
    records = AttendanceRecord.objects.filter(
        date__range=[start_date, end_date]
    )
    
    # Apply filters if provided
    if section_id:
        records = records.filter(student__section_id=section_id)
    
    if course_id:
        records = records.filter(student__section__course_id=course_id)
    
    if year_level:
        records = records.filter(student__year_level=year_level)
    
    if search_query:
        records = records.filter(
            Q(student__student_id__icontains=search_query) |
            Q(student__first_name__icontains=search_query) |
            Q(student__last_name__icontains=search_query)   # <--- FIXED
        )
    
    # Calculate summary data - ALL records within the date range
    summary = {
        'am_in': records.exclude(morning_in__isnull=True).count(),
        'am_out': records.exclude(morning_out__isnull=True).count(),
        'pm_in': records.exclude(afternoon_in__isnull=True).count(),
        'pm_out': records.exclude(afternoon_out__isnull=True).count(),
        'total_records': records.count(),
        'total_students': records.values('student').distinct().count(),
    }
    
    # Calculate daily data for the entire date range
    daily_data = []
    current_date = start_date
    while current_date <= end_date:
        date_records = records.filter(date=current_date)
        
        # Get unique students who had any attendance on this date
        students_with_attendance = date_records.exclude(
            Q(morning_in__isnull=True) & 
            Q(afternoon_in__isnull=True)
        ).values('student').distinct().count()
        
        daily_data.append({
            'date': current_date.strftime('%b %d'),
            'full_date': current_date.strftime('%Y-%m-%d'),
            'am_in': date_records.exclude(morning_in__isnull=True).count(),
            'am_out': date_records.exclude(morning_out__isnull=True).count(),
            'pm_in': date_records.exclude(afternoon_in__isnull=True).count(),
            'pm_out': date_records.exclude(afternoon_out__isnull=True).count(),
            'total': students_with_attendance,
            'am_total': date_records.exclude(morning_in__isnull=True).count(),
            'pm_total': date_records.exclude(afternoon_in__isnull=True).count()
        })
        current_date += timedelta(days=1)
    
    return JsonResponse({
        'summary': summary,
        'daily': daily_data,
        'date_range': {
            'start': start_date.strftime('%Y-%m-%d'),
            'end': end_date.strftime('%Y-%m-%d')
        }
    })

@login_required
def export_attendance(request):
    # Get the same filters as the reports view
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    section_id = request.GET.get('section')
    course_id = request.GET.get('course')
    year_level = request.GET.get('year_level')
    search_query = request.GET.get('search', '')
    
    # If no dates provided, export ALL records (remove the 7-day default)
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        # Set to a very old date to get all records
        start_date = date(2000, 1, 1)  # Or your system's start date
    
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        # Set to future date to get all records up to today and beyond
        end_date = date.today() + timedelta(days=365)  # Include future dates if any
    
    # Get all records with the same filters
    records = AttendanceRecord.objects.filter(
        date__range=[start_date, end_date]
    ).select_related('student', 'student__section', 'student__section__course')
    
    if section_id:
        records = records.filter(student__section_id=section_id)
    
    if course_id:
        records = records.filter(student__section__course_id=course_id)
    
    if year_level:
        records = records.filter(student__year_level=year_level)
    
    if search_query:
        records = records.filter(
            Q(student__student_id__icontains=search_query) |
            Q(student__first_name__icontains=search_query) |
            Q(student__last_name__icontains=search_query)
        )
    
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Records"
    
    # Add headers
    headers = [
        "Date", "Student ID", "Student Name", "Course", "Section", 
        "Morning In", "Morning In Photo", "Morning Out", "Morning Out Photo",
        "Afternoon In", "Afternoon In Photo", "Afternoon Out", "Afternoon Out Photo"
    ]
    
    # Write headers with bold style
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}1"]
        cell.value = header
        cell.font = Font(bold=True)
    
    # Function to check if photo URL is accessible
    def is_photo_accessible(url):
        if not url:
            return False
        
        # Check if it's a local file path
        if url.startswith('/') or url.startswith(settings.MEDIA_URL):
            # Extract the relative path from the URL
            if url.startswith(settings.MEDIA_URL):
                relative_path = url.replace(settings.MEDIA_URL, '')
            else:
                relative_path = url.replace('/', '', 1) if url.startswith('/') else url
            
            # Check if file exists
            file_path = os.path.join(settings.MEDIA_ROOT, relative_path)
            return os.path.exists(file_path)
        
        # For external URLs (like Google Drive), we'll assume they're accessible
        return True
    
    # Function to get photo status
    def get_photo_status(url):
        if not url:
            return "No photo"
        
        if is_photo_accessible(url):
            return url  # Return the URL if accessible
        else:
            return "Photo not found or moved"
    
    # Write attendance data
    for row_num, record in enumerate(records, 2):
        ws.cell(row=row_num, column=1, value=record.date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=2, value=record.student.student_id)
        ws.cell(row=row_num, column=3, value=record.student.full_name)
        ws.cell(row=row_num, column=4, value=record.student.section.course.name if record.student.section else "")
        ws.cell(row=row_num, column=5, value=record.student.section.name if record.student.section else "")
        
        # Morning In with photo link
        ws.cell(row=row_num, column=6, value=record.morning_in.strftime('%H:%M:%S') if record.morning_in else "")
        ws.cell(row=row_num, column=7, value=get_photo_status(record.morning_in_photo_url))
        
        # Morning Out with photo link
        ws.cell(row=row_num, column=8, value=record.morning_out.strftime('%H:%M:%S') if record.morning_out else "")
        ws.cell(row=row_num, column=9, value=get_photo_status(record.morning_out_photo_url))
        
        # Afternoon In with photo link
        ws.cell(row=row_num, column=10, value=record.afternoon_in.strftime('%H:%M:%S') if record.afternoon_in else "")
        ws.cell(row=row_num, column=11, value=get_photo_status(record.afternoon_in_photo_url))
        
        # Afternoon Out with photo link
        ws.cell(row=row_num, column=12, value=record.afternoon_out.strftime('%H:%M:%S') if record.afternoon_out else "")
        ws.cell(row=row_num, column=13, value=get_photo_status(record.afternoon_out_photo_url))
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Generate filename with date range info
    if start_date and end_date:
        filename = f"attendance_export_{start_date}_to_{end_date}.xlsx"
    else:
        filename = f"attendance_export_all_records.xlsx"
    
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    
    return response


@csrf_exempt
def process_pending_emails(request):
    if request.method == 'POST':
        try:
            pending_records = AttendanceRecord.objects.filter(
                email_sent=False,
                email_attempted=False
            ).select_related('student')
            
            processed = 0
            
            for record in pending_records:
                try:
                    record.email_attempted = True
                    record.save()
                    
                    # Determine email type and get the corresponding photo URL
                    email_type = None
                    photo_url = None
                    photo_filename = None
                    
                    if record.morning_in and not record.morning_out:
                        email_type = 'time_in'
                        photo_url = record.morning_in_photo_url
                        photo_filename = f"{record.student.student_id}_morning_in.jpg"
                    elif record.morning_out:
                        email_type = 'time_out'
                        photo_url = record.morning_out_photo_url
                        photo_filename = f"{record.student.student_id}_morning_out.jpg"
                    elif record.afternoon_in and not record.afternoon_out:
                        email_type = 'time_in'
                        photo_url = record.afternoon_in_photo_url
                        photo_filename = f"{record.student.student_id}_afternoon_in.jpg"
                    elif record.afternoon_out:
                        email_type = 'time_out'
                        photo_url = record.afternoon_out_photo_url
                        photo_filename = f"{record.student.student_id}_afternoon_out.jpg"
                    else:
                        continue
                    
                    # Download the photo from local storage if available
                    photo_file = None
                    if photo_url and photo_url.startswith(settings.MEDIA_URL):
                        try:
                            # Extract the relative path from the URL
                            relative_path = photo_url.replace(settings.MEDIA_URL, '')
                            full_path = os.path.join(settings.MEDIA_ROOT, relative_path)
                            
                            # Open the file
                            photo_file = open(full_path, 'rb')
                            logger.info(f"Loaded local photo for pending email: {full_path}")
                            
                        except Exception as e:
                            logger.error(f"Error loading local photo {photo_url}: {str(e)}")
                            photo_file = None
                    
                    # Send email with photo if available
                    success = send_attendance_email(
                        student=record.student,
                        record=record,
                        email_type=email_type,
                        photo_file=photo_file,  # Pass the file object
                        photo_filename=photo_filename
                    )
                    
                    # Close the file if it was opened
                    if photo_file:
                        photo_file.close()
                    
                    if success:
                        record.email_sent = True
                        record.save()
                        processed += 1
                        logger.info(f"Successfully sent pending email for {record.student.full_name}")
                    else:
                        logger.warning(f"Failed to send pending email for {record.student.full_name}")
                        
                except Exception as e:
                    logger.error(f"Error processing pending email for {record.student.full_name}: {str(e)}")
                    continue
            
            return JsonResponse({
                'success': True,
                'processed': processed,
                'total_pending': pending_records.count()
            })
            
        except Exception as e:
            logger.error(f"Error in process_pending_emails: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    
    return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)

@login_required
def google_drive_auth(request):
    """Start Google Drive OAuth authentication"""
    try:
        auth_url = drive_oauth_service.get_authorization_url()
        if auth_url:
            return redirect(auth_url)
        else:
            messages.error(request, "Failed to initialize Google Drive authentication")
            return redirect('system_settings')
    except Exception as e:
        logger.error(f"Error in Google Drive auth: {e}")
        messages.error(request, f"Authentication error: {str(e)}")
        return redirect('system_settings')

@csrf_exempt
def google_drive_callback(request):
    """Handle Google OAuth callback"""
    try:
        if 'code' in request.GET:
            code = request.GET['code']
            success = drive_oauth_service.save_credentials(code)
            
            if success:
                messages.success(request, "Google Drive authentication successful!")
                # Reinitialize the service
                drive_oauth_service.initialize_service()
            else:
                messages.error(request, "Google Drive authentication failed")
        else:
            messages.error(request, "Google Drive authentication was cancelled")
        
        return redirect('system_settings')
    except Exception as e:
        logger.error(f"Error in Google Drive callback: {e}")
        messages.error(request, f"Authentication error: {str(e)}")
        return redirect('system_settings')

@login_required
def google_drive_status(request):
    """Check Google Drive authentication status"""
    try:
        return JsonResponse({
            'authenticated': drive_oauth_service.is_authenticated()
        })
    except Exception as e:
        logger.error(f"Error checking Google Drive status: {e}")
        return JsonResponse({
            'authenticated': False,
            'error': str(e)
        })

@login_required
def check_drive_status(request):
    """Force check Google Drive status"""
    drive_oauth_service.initialize_service()
    return JsonResponse({
        'authenticated': drive_oauth_service.is_authenticated(),
        'has_credentials': drive_oauth_service.credentials is not None,
        'has_service': drive_oauth_service.service is not None
    })

@login_required
def debug_drive_cache(request):
    """Debug endpoint to check cache status"""
    from django.core.cache import cache
    
    cached_creds = cache.get('google_drive_credentials')
    drive_oauth_service.initialize_service()
    
    return JsonResponse({
        'cached_credentials_exists': cached_creds is not None,
        'cached_credentials_keys': list(cached_creds.keys()) if cached_creds else [],
        'has_credentials': drive_oauth_service.credentials is not None,
        'has_service': drive_oauth_service.service is not None,
        'is_authenticated': drive_oauth_service.is_authenticated(),
        'cache_keys': list(cache.__dict__.keys()) if hasattr(cache, '__dict__') else 'N/A'
    })

# Add this to your views.py file
@login_required
def debug_drive_auth(request):
    """Debug view to check Google Drive authentication status"""
    drive_oauth_service.initialize_service()
    
    # Check cache directly
    from django.core.cache import cache
    cached_creds = cache.get('google_drive_credentials')
    
    debug_info = {
        'cached_credentials_exists': cached_creds is not None,
        'has_credentials': drive_oauth_service.credentials is not None,
        'has_service': drive_oauth_service.service is not None,
        'is_authenticated': drive_oauth_service.is_authenticated(),
        'client_id_configured': bool(settings.GOOGLE_OAUTH2_CLIENT_ID),
        'client_secret_configured': bool(settings.GOOGLE_OAUTH2_CLIENT_SECRET),
        'redirect_uri_configured': bool(settings.GOOGLE_OAUTH2_REDIRECT_URI),
    }
    
    return JsonResponse(debug_info)

# views.py - Add this view
@login_required
def google_drive_disconnect(request):
    """Disconnect Google Drive and clear credentials"""
    try:
        drive_oauth_service.clear_credentials()
        messages.success(request, "Google Drive disconnected successfully")
    except Exception as e:
        logger.error(f"Error disconnecting Google Drive: {e}")
        messages.error(request, "Error disconnecting Google Drive")
    
    return redirect('system_settings')

@login_required
def clear_import_errors(request):
    if 'import_errors' in request.session:
        del request.session['import_errors']
    return JsonResponse({'success': True})

@login_required
def print_attendance_report(request):
    """
    Generates a printer-friendly HTML page for attendance records
    """
    # 1. Capture Filters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    institute_id = request.GET.get('institute')
    section_id = request.GET.get('section')
    course_id = request.GET.get('course')
    year_level = request.GET.get('year_level')
    
    # 2. Date Handling
    if start_date:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date_obj = date(2000, 1, 1)
    
    if end_date:
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        end_date_obj = date.today() + timedelta(days=365)
    
    # 3. Query Data
    records = AttendanceRecord.objects.filter(
        date__range=[start_date_obj, end_date_obj]
    ).select_related(
        'student', 
        'student__section', 
        'student__section__course',
        'student__section__course__institute'
    )
    
    # 4. Apply Filters in correct order (institute -> course -> year -> section)
    filter_description = []
    
    # Institute filter
    if institute_id:
        records = records.filter(student__section__course__institute_id=institute_id)
        try:
            institute = Institute.objects.get(id=institute_id)
            filter_description.append(f"Department: {institute.code}")
        except Institute.DoesNotExist:
            pass
    
    # Course filter
    if course_id:
        records = records.filter(student__section__course_id=course_id)
        try:
            course = Course.objects.get(id=course_id)
            filter_description.append(f"Program: {course.code}")
        except Course.DoesNotExist:
            pass
    
    # Year level filter
    if year_level:
        records = records.filter(student__year_level=year_level)
        filter_description.append(f"Year: {year_level}")
    
    # Section filter (most specific)
    if section_id:
        records = records.filter(student__section_id=section_id)
        try:
            section = Section.objects.get(id=section_id)
            filter_description.append(f"Section: {section.name}")
        except Section.DoesNotExist:
            pass
    
    # 5. Sorting (Date desc, then Time desc)
    # Convert to list to sort with python logic like the main view
    records = list(records)
    
    def get_latest_time(record):
        times = [t for t in [record.morning_in, record.morning_out, 
                            record.afternoon_in, record.afternoon_out] if t]
        return max(times) if times else datetime.min.time()
    
    records.sort(key=lambda r: (r.date, get_latest_time(r)), reverse=True)
    
    # 6. Calculate statistics for the report
    total_records = len(records)
    total_students = len(set(r.student.id for r in records))
    
    # Count morning/afternoon attendance
    morning_in_count = sum(1 for r in records if r.morning_in)
    morning_out_count = sum(1 for r in records if r.morning_out)
    afternoon_in_count = sum(1 for r in records if r.afternoon_in)
    afternoon_out_count = sum(1 for r in records if r.afternoon_out)
    
    # 7. Context for Template
    context = {
        'records': records,
        'start_date': start_date_obj if start_date else None,
        'end_date': end_date_obj if end_date else date.today(),
        'filters': ", ".join(filter_description) if filter_description else "All Records",
        'generated_at': datetime.now(),
        'generated_by': request.user.username if request.user.is_authenticated else "System",
        'total_records': total_records,
        'total_students': total_students,
        'morning_in_count': morning_in_count,
        'morning_out_count': morning_out_count,
        'afternoon_in_count': afternoon_in_count,
        'afternoon_out_count': afternoon_out_count,
    }
    
    return render(request, 'attendance_app/admin/print_report.html', context)

def forgot_password_view(request):
    """View for requesting password reset"""
    if request.user.is_authenticated:
        return redirect('admin_dashboard')
    
    if request.method == 'POST':
        form = ForgotPasswordForm(request.POST)
        if form.is_valid():
            try:
                user = form.get_user()
                
                # Generate and send OTP
                otp_record = generate_and_send_otp(user)
                
                if otp_record:
                    # Store user ID in session for OTP verification
                    request.session['reset_user_id'] = user.id
                    request.session['reset_attempts'] = 0  # Track OTP attempts
                    
                    messages.success(
                        request, 
                        f"An OTP has been sent to your email address. Please check your inbox and spam folder."
                    )
                    return redirect('verify_otp')
                else:
                    messages.error(request, "Failed to send OTP. Please try again.")
            except Exception as e:
                logger.error(f"Error in forgot password: {e}")
                messages.error(request, "An error occurred. Please try again.")
    else:
        form = ForgotPasswordForm()
    
    return render(request, 'attendance_app/forgot_password.html', {'form': form})

# attendance_app/views.py

def verify_otp_view(request):
    """View for verifying OTP"""
    if request.user.is_authenticated:
        return redirect('admin_dashboard')
    
    # Check if user is in password reset flow
    user_id = request.session.get('reset_user_id')
    if not user_id:
        messages.error(request, "Please request a password reset first.")
        return redirect('forgot_password')
    
    try:
        from django.contrib.auth import get_user_model
        User = get_user_model()
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        messages.error(request, "Invalid session. Please request a new password reset.")
        del request.session['reset_user_id']
        return redirect('forgot_password')
    
    # Check attempt limit
    attempts = request.session.get('reset_attempts', 0)
    if attempts >= 5:
        messages.error(request, "Too many failed attempts. Please request a new OTP.")
        del request.session['reset_user_id']
        del request.session['reset_attempts']
        return redirect('forgot_password')
    
    # NEW: Calculate remaining seconds for the timer
    remaining_seconds = 0
    latest_otp = PasswordResetOTP.objects.filter(user=user, is_used=False).order_by('-created_at').first()
    
    if latest_otp:
        now = timezone.now()
        if latest_otp.expires_at > now:
            delta = latest_otp.expires_at - now
            remaining_seconds = int(delta.total_seconds())
        else:
            remaining_seconds = 0

    if request.method == 'POST':
        if 'resend_otp' in request.POST:
            # Resend OTP
            otp_record = generate_and_send_otp(user)
            if otp_record:
                messages.success(request, "A new OTP has been sent to your email.")
                # Update remaining seconds for the redirect
                remaining_seconds = 60 
            else:
                messages.error(request, "Failed to resend OTP. Please try again.")
            
            # Redirect to self to refresh the page/timer cleanly
            return redirect('verify_otp')
        
        form = VerifyOTPForm(request.POST, user=user)
        if form.is_valid():
            # Get the verified OTP code
            verified_otp = form.cleaned_data['otp_code']
            
            # Store the verified OTP in session
            request.session['verified_otp'] = verified_otp
            request.session['otp_verified'] = True
            
            messages.success(request, "OTP verified successfully. Please set your new password.")
            return redirect('reset_password')
        else:
            # Increment attempt counter
            request.session['reset_attempts'] = attempts + 1
            if attempts + 1 >= 5:
                messages.error(request, "Too many failed attempts. Please request a new OTP.")
                del request.session['reset_user_id']
                del request.session['reset_attempts']
                return redirect('forgot_password')
    else:
        form = VerifyOTPForm(user=user)
    
    return render(request, 'attendance_app/verify_otp.html', {
        'form': form,
        'user': user,
        'attempts_remaining': 5 - attempts,
        'remaining_seconds': remaining_seconds # Pass this to template
    })

# Update in views.py - reset_password_view function
def reset_password_view(request):
    """View for resetting password after OTP verification"""
    if request.user.is_authenticated:
        return redirect('admin_dashboard')
    
    # Check if OTP was verified
    if not request.session.get('otp_verified'):
        messages.error(request, "Please verify OTP first.")
        return redirect('forgot_password')
    
    # Get the verified OTP from session
    verified_otp = request.session.get('verified_otp')
    if not verified_otp:
        messages.error(request, "OTP session expired. Please request a new password reset.")
        return redirect('forgot_password')
    
    user_id = request.session.get('reset_user_id')
    if not user_id:
        messages.error(request, "Session expired. Please request a new password reset.")
        return redirect('forgot_password')
    
    try:
        from django.contrib.auth import get_user_model
        User = get_user_model()
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        messages.error(request, "Invalid session. Please request a new password reset.")
        del request.session['reset_user_id']
        del request.session['otp_verified']
        del request.session['verified_otp']
        return redirect('forgot_password')
    
    if request.method == 'POST':
        form = ResetPasswordForm(request.POST)
        if form.is_valid():
            # Use the verified OTP from session instead of form
            new_password = form.cleaned_data['new_password']
            
            # Verify OTP and reset password
            success, message = verify_otp_and_reset_password(user, verified_otp, new_password)
            
            if success:
                # Clear all session data
                keys_to_delete = ['reset_user_id', 'otp_verified', 'verified_otp', 'reset_attempts']
                for key in keys_to_delete:
                    if key in request.session:
                        del request.session[key]
                
                messages.success(request, "Password reset successfully! You can now login with your new password.")
                return redirect('login')
            else:
                messages.error(request, message)
        else:
            # Check if form has specific errors
            if 'new_password' in form.errors:
                messages.error(request, form.errors['new_password'][0])
            elif 'confirm_password' in form.errors:
                messages.error(request, form.errors['confirm_password'][0])
    else:
        form = ResetPasswordForm()
    
    return render(request, 'attendance_app/reset_password.html', {
        'form': form,
        'user': user
    })

# In views.py - update the print_students_report function
@login_required
def print_students_report(request):
    """
    Generates a printer-friendly HTML page for student records
    """
    # 1. Capture Filters
    search_query = request.GET.get('search', '')
    institute_id = request.GET.get('institute', '')
    course_id = request.GET.get('course', '')
    year_level = request.GET.get('year_level', '')
    section_id = request.GET.get('section', '')
    sort_by = request.GET.get('sort', 'name_asc')
    
    # 2. Query Data
    students = Student.objects.select_related('section__course__institute')
    
    # Apply filters
    filter_description = []
    
    if search_query:
        students = students.filter(
            Q(student_id__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query)
        )
        filter_description.append(f"Search: '{search_query}'")
    
    if institute_id:
        students = students.filter(section__course__institute_id=institute_id)
        try:
            institute = Institute.objects.get(id=institute_id)
            filter_description.append(f"Department: {institute.code}")
        except Institute.DoesNotExist:
            pass
    
    if course_id:
        students = students.filter(section__course_id=course_id)
        try:
            course = Course.objects.get(id=course_id)
            filter_description.append(f"Program: {course.code}")
        except Course.DoesNotExist:
            pass
    
    if year_level:
        students = students.filter(year_level=year_level)
        filter_description.append(f"Year Level: {year_level}")
    
    if section_id:
        students = students.filter(section_id=section_id)
        try:
            section = Section.objects.get(id=section_id)
            filter_description.append(f"Section: {section.name}")
        except Section.DoesNotExist:
            pass
    
    # Sorting - UPDATED WITH DATE SORTING
    if sort_by == 'name_asc':
        students = students.order_by('last_name', 'first_name')
    elif sort_by == 'name_desc':
        students = students.order_by('-last_name', '-first_name')
    elif sort_by == 'date_created_asc':
        students = students.order_by('created_at')
    elif sort_by == 'date_created_desc':
        students = students.order_by('-created_at')
    elif sort_by == 'date_modified_asc':
        students = students.order_by('updated_at')
    elif sort_by == 'date_modified_desc':
        students = students.order_by('-updated_at')
    elif sort_by == 'date_asc':
        students = students.order_by('id')
    elif sort_by == 'date_desc':
        students = students.order_by('-id')
    
    # Convert to list for template
    students_list = list(students)
    
    # 3. Context for Template
    context = {
        'students': students_list,
        'filters': ", ".join(filter_description) if filter_description else "All Students",
        'generated_at': datetime.now(),
        'generated_by': request.user.username if request.user.is_authenticated else "System",
        'total_count': len(students_list)
    }
    
    return render(request, 'attendance_app/admin/print_students.html', context)

# API Views
def get_sections_by_year(request, year_level):
    sections = Section.objects.filter(year_level=year_level).values('id', 'name')
    return JsonResponse(list(sections), safe=False)